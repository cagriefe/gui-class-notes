import ttkbootstrap as ttk  # Importing ttkbootstrap for themed tkinter widgets
from tkinter import filedialog as fd  # Importing filedialog for file operations
from ttkbootstrap.dialogs import Messagebox as msg  # Importing Messagebox for displaying messages
from openpyxl import load_workbook  # Importing load_workbook for reading Excel files
from openpyxl import Workbook  # Importing Workbook for creating Excel files
import sqlite3
import matplotlib.pyplot as plt

class ToDoDatabase:

    def __init__(self, db_name="todo_list.db"):
        self.db_name = db_name

    def create_table(self):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("""
            create table if not exists todolist (
                id integer primary key autoincrement,
                item text not null,
                status text default 'Pending'
            );
        """)
        conn.commit()
        conn.close()

    def add_item(self, item):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("insert into todolist (item) values (?)", (item,))
        conn.commit()
        conn.close()

    def get_items(self):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("select id, item, status from todolist")
        items = cur.fetchall()
        conn.close()

        return items

    def delete_item(self, item_id):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("delete from todolist where id = ?", (item_id,))
        conn.commit()
        conn.close()

    def update_status(self, item_id, status):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("update todolist set status = ? where id = ?", 
                    (status, item_id))
        conn.commit()
        conn.close()


class ToDoList(ttk.Window):  # Defining a class that inherits from ttk.Window
    def __init__(self):  # Constructor method
        super().__init__(themename="litera")  # Calling the constructor of the parent class with a theme
        self.geometry("600x720")  # Setting the window size
        self.title("To-Do List App")  # Setting the window title
        self.db = ToDoDatabase()  # Creating an instance of the database
        self.status_var = ttk.IntVar()  # Creating an IntVar for the status checkbutton
        self.create_widgets()  # Calling method to create widgets
        self.create_layout()  # Calling method to create layout
        self.mainloop()  # Starting the main event loop

    def create_widgets(self):  # Method to create widgets
        # Widgets for adding new To-Do items
        self.frm_add = ttk.Labelframe(self, text="Add To-Do Item")  # Creating a Labelframe for adding items
        self.entry_item = ttk.Entry(self.frm_add)  # Creating an entry widget for the item
        self.btn_add = ttk.Button(self.frm_add, text="Add", command=self.add_item)  # Creating a button to add the item

        # Widgets for listing To-Do items
        self.frm_list = ttk.Labelframe(self, text="To-Do List")  # Creating a Labelframe for the list
        self.tv = ttk.Treeview(self.frm_list, columns=("Item", "Status"),
                               show="headings", selectmode="browse")  # Creating a Treeview widget
        self.tv.heading("Item", text="Item")  # Setting the heading for the Item column
        self.tv.heading("Status", text="Status")  # Setting the heading for the Status column

        # Widgets for deleting/updating To-Do items
        self.frm_actions = ttk.Labelframe(self, text="Actions")  # Creating a Labelframe for actions
        self.btn_delete = ttk.Button(self.frm_actions, text="Delete", command=self.delete_item)  # Creating a button to delete the item
        self.cbtn_status = ttk.Checkbutton(self.frm_actions, text="Mark as Completed",
                                           variable=self.status_var,
                                           command=self.toggle_status, bootstyle="round-toggle")  # Creating a checkbutton to mark the item as completed

        # Event bindings for the Treeview and Entry widgets
        self.tv.bind("<<TreeviewSelect>>", self.on_item_select)  # Binding the Treeview selection event
        self.entry_item.bind("<Return>", lambda e : self.add_item())  # Binding the Return key to the add_item method
        
        # Event bindings for the Import / Export functionalities
        self.bind("<Control-i>", lambda e : self.import_data())  # Binding Ctrl+I to the import_data method
        self.bind("<Control-e>", lambda e : self.export_data())  # Binding Ctrl+E to the export_data method

        # Event binding for pie chart view
        self.bind("<Control-Shift-C>", lambda e : self.show_pie_chart())
        
        # Populate the Treeview
        self.refresh_list()  # Calling method to populate the Treeview

    def create_layout(self):  # Method to create layout
        # Place Labelframe widgets
        self.frm_add.pack(fill="x", padx=15, pady=(15, 0))  # Packing the add item frame with padding
        self.frm_list.pack(fill="both", expand=True, padx=15, pady=(15, 0))  # Packing the list frame with padding
        self.frm_actions.pack(fill="x", padx=15, pady=15)  # Packing the actions frame with padding

        # Place entry and add button widgets
        self.entry_item.pack(side="left", padx=(15, 0), pady=15,
                             fill="x", expand=True)  # Packing the entry widget with padding
        self.btn_add.pack(side="right", padx=15, pady=15)  # Packing the add button with padding

        # Place the Treeview widget
        self.tv.pack(fill="both", expand=True, padx=15, pady=15)  # Packing the Treeview with padding

        # Place delete button and status checkbutton widgets
        self.btn_delete.pack(side="left", padx=15, pady=15)  # Packing the delete button with padding
        self.cbtn_status.pack(side="right", padx=15, pady=15)  # Packing the status checkbutton with padding

    def enable_actions(self):  # Method to enable action buttons
        self.btn_delete.configure(state="normal")  # Enabling the delete button
        self.cbtn_status.configure(state="normal")  # Enabling the status checkbutton

    def disable_actions(self):  # Method to disable action buttons
        self.btn_delete.configure(state="disabled")  # Disabling the delete button
        self.cbtn_status.configure(state="disabled")  # Disabling the status checkbutton

    def on_item_select(self, event):  # Method to handle item selection
        selected_item = self.tv.selection()  # Getting the selected item
        if selected_item:  # If an item is selected
            item_id = int(selected_item[0])  # Getting the item ID
            item_status = self.tv.item(item_id)["values"][1]  # Getting the item status
            if item_status == "Done":  # If the item is marked as done
                self.status_var.set(1)  # Setting the checkbutton to checked
            else:  # If the item is not marked as done
                self.status_var.set(0)  # Setting the checkbutton to unchecked

    def refresh_list(self):  # Method to refresh the list
        for item in self.tv.get_children():  # Clearing the Treeview content
            self.tv.delete(item)  # Deleting each item

        for row in self.db.get_items():  # Fetching updated data and populating the Treeview
            self.tv.insert(parent="", index="end", iid=row[0], values=(row[1], row[2]))  
                                                                    # Inserting each row
        tv_items = self.tv.get_children()  # Getting the Treeview items
        if tv_items:  # If there are items in the Treeview
            self.enable_actions()  # Enabling the action buttons
            self.tv.selection_set(tv_items[-1])  # Selecting the last item
        else:  # If there are no items in the Treeview
            self.disable_actions()  # Disabling the action buttons

    def add_item(self):  # Method to add an item
        item = self.entry_item.get().strip()  # Getting the item from the entry widget
        if item:  # If the item is not empty
            self.db.add_item(item)  # Adding the item to the database
            self.refresh_list()  # Refreshing the list
            self.entry_item.delete(0, "end")  # Clearing the entry widget
            self.entry_item.focus_set()  # Setting focus to the entry widget



    def delete_item(self):  # Method to delete an item
        selected_item = self.tv.selection()  # Getting the selected item
        if selected_item:  # If an item is selected
            item_id = int(selected_item[0])  # Getting the item ID
            self.db.delete_item(item_id)  # Deleting the item from the database
            self.refresh_list()  # Refreshing the list

    def toggle_status(self):  # Method to toggle the status of an item
        selected_item = self.tv.selection()  # Getting the selected item
        if selected_item:  # If an item is selected
            item_id = int(selected_item[0])  # Getting the item ID
            if self.status_var.get() == 1:  # If the checkbutton is checked
                new_status = "Done"  # Setting the new status to Done
            else:  # If the checkbutton is unchecked
                new_status = "Pending"  # Setting the new status to Pending
            self.db.update_status(item_id, new_status)  # Updating the status in the database
            self.refresh_list()  # Refreshing the list
            self.tv.selection_set(selected_item)  # Keeping the current item selected

    def import_data(self):  # Method to import data
        file_filter = (("Excel files", "*.xlsx"),  # Setting the file filter for Excel files
                       ("All files", "*.*"))  # Setting the file filter for all files
        source_file = fd.askopenfile(title="Choose file",
                                     filetypes=file_filter)  # Opening a file dialog to choose the source file

        if source_file is not None:  # If a file is selected
            wb = load_workbook(source_file.name)  # Loading the source file
            ws1 = wb.active  # Getting the active worksheet
            skip_first = True  # Flag to skip the header row

            for row in ws1.iter_rows(values_only=True):  # Reading the file content
                if skip_first:  # If it's the header row
                    skip_first = False  # Skip the header row
                    continue

                item, status = row[0], row[1]  # Getting the first two columns
                if item:  # If the item is not empty
                    self.db.add_item(item)  # Inserting the item to the database

                    if status == "Done":  # If the status is "Done"
                        self.db.update_status(self.db.get_items()[-1][0], "Done")  # Updating the status

            self.refresh_list()  # Re-populating the Treeview


    def export_data(self):  # Method to export data
        if not self.tv.get_children():  # If there are no items in the Treeview
            msg.show_error(title="Export Data", message="No data available")  # Showing an error message
            return

        wb = Workbook()  # Creating a new workbook
        ws = wb.active  # Getting the active worksheet

        ws.cell(row=1, column=1, value="Item")  # Adding column header for Item
        ws.cell(row=1, column=2, value="Status")  # Adding column header for Status

        row_count = 2  # Starting row count for data
        for item_id in self.tv.get_children():  # Reading data from Treeview
            ws.cell(row=row_count, column=1,
                    value=self.tv.item(item_id)["values"][0])  # Adding item to the worksheet
            ws.cell(row=row_count, column=2, 
                    value=self.tv.item(item_id)["values"][1])  # Adding status to the worksheet
            row_count += 1  # Incrementing row count

        total_items_row = row_count + 1  # Row for total items
        completed_items_row = row_count + 2  # Row for completed items
        not_completed_items_row = row_count + 3  # Row for not completed items

        ws.cell(row=total_items_row, column=1, value="Total items:")  # Adding label for total items
        ws.cell(row=total_items_row, column=2, 
                value=f"=COUNTA(A2:A{row_count - 1})")  # Adding formula for total items

        ws.cell(row=completed_items_row, column=1,
                value="Completed items:")  # Adding label for completed items
        ws.cell(row=completed_items_row, column=2, 
                value=f"=COUNTIF(B2:B{row_count - 1}, \"Done\")")  # Adding formula for completed items

        ws.cell(row=not_completed_items_row, column=1, 
                value="Not completed items:")  # Adding label for not completed items
        ws.cell(row=not_completed_items_row, column=2, 
                value=f"=COUNTIF(B2:B{row_count - 1}, \"Pending\")")  # Adding formula for not completed items

        wb.save("exported_list.xlsx")  # Saving the workbook
        msg.show_info(title="Export Data",
                      message="exported and saved as 'exported_list.xlsx'.")  # Showing an info message
    
    def show_pie_chart(self):
        # Retrieve items from the database
        items = self.db.get_items()
        # If there are no items, exit the function
        if len(items) == 0:
            return 
        # Count the number of completed items
        completed = sum(1 for item in items if item[2] == "Done")
        # Count the number of not completed items
        not_completed = sum(1 for item in items if item[2] == "Pending")
        # Define labels for the pie chart
        labels = ["Completed", "Not Completed"]
        # Define sizes for each section of the pie chart
        sizes = [completed, not_completed]
        # Define colors for each section of the pie chart
        colors = ["lightblue", "red"]
        # Define the explode parameter to highlight the "Completed" section
        explode = [0.1, 0] # Highlight the "Completed" section
        # Create the pie chart
        plt.pie(sizes, explode=explode, labels=labels, 
                colors=colors, autopct="%.1f%%", shadow=True, startangle=90)
        # Set the title of the pie chart
        plt.title("To-Do list Status")
        # Display the pie chart
        plt.show()

gradebook_db = ToDoDatabase()
gradebook_db.create_table()

app = ToDoList()  # Creating an instance of ToDoList

