import ttkbootstrap as ttk
from tkinter import messagebox as msg
import sqlite3
from openpyxl import load_workbook, workbook

class LibraryDatabase:
    def __init__(self, db_name="library.db"):
        self.db_name = db_name
        
    def create_table(self):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("""
                    CREATE TABLE IF NOT EXISTS books (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        title TEXT,
                        author TEXT,
                        year INTEGER
                    )
                    """)
        conn.commit()
        conn.close()
    
    def add_book(self, title, author, year):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("INSERT INTO books (title, author, year) VALUES (?, ?, ?)", (title, author, year))
        conn.commit()
        conn.close()
    
    def get_items(self):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("SELECT id, title, author, year FROM books")
        items = cur.fetchall()
        conn.close()
        return items

    def delete_book(self, book_id):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("DELETE FROM books WHERE id = ?", (book_id,))
        conn.commit()
        conn.close()

    def update_book(self, book_id, title, author, year):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("UPDATE books SET title = ?, author = ?, year = ? WHERE id = ?", (title, author, year, book_id))
        conn.commit()
        conn.close()

class BookManagement(ttk.Window):
    def __init__(self):
        super().__init__()
        self.geometry("800x600")
        self.title("Library Book Management System")
        self.db = LibraryDatabase()
        self.db.create_table()
        self.create_widgets()
        self.create_layout()
        self.refresh_list()
        self.mainloop()
    
    def create_widgets(self):
        self.lbl_title = ttk.Label(self, text="Book Title")
        self.lbl_author = ttk.Label(self, text="Author")
        self.lbl_year = ttk.Label(self, text="Year")
        
        self.txt_title = ttk.Entry(self)
        self.txt_author = ttk.Entry(self)
        self.txt_year = ttk.Entry(self)
        
        self.btn_add = ttk.Button(self, text="Add Book", command=self.add_book)
        self.btn_clear = ttk.Button(self, text="Clear Fields", command=self.clear_fields)
        
        self.tv = ttk.Treeview(self, columns=("Book Title", "Author", "Year"), show="headings", selectmode="browse")
        self.tv.heading("Book Title", text="Book Title")
        self.tv.heading("Author", text="Author")
        self.tv.heading("Year", text="Year")
        
        self.btn_update = ttk.Button(self, text="Update Book", command=self.update_book)
        self.btn_delete = ttk.Button(self, text="Delete Book", command=self.delete_book)
        
        self.bind("<F1>", lambda _: self.db.create_table())
        self.bind("<F2>", lambda _: self.export_data())
        
    def create_layout(self):
        self.columnconfigure(index=0, weight=1, uniform="eq")
        self.columnconfigure(index=1, weight=1, uniform="eq")
        self.rowconfigure(index=0, weight=1, uniform="eq")
        self.rowconfigure(index=1, weight=1, uniform="eq")
        self.rowconfigure(index=2, weight=1, uniform="eq")
        self.rowconfigure(index=3, weight=1, uniform="eq")
        self.rowconfigure(index=4, weight=6, uniform="eq")
        self.rowconfigure(index=5, weight=1, uniform="eq")
        
        self.lbl_title.grid(column=0, row=0)
        self.txt_title.grid(column=1, row=0)
        
        self.lbl_author.grid(column=0, row=1)
        self.txt_author.grid(column=1, row=1)
        
        self.lbl_year.grid(column=0, row=2)
        self.txt_year.grid(column=1, row=2)
        
        self.btn_add.grid(column=0, row=3)
        self.btn_clear.grid(column=1, row=3)
        
        self.tv.grid(column=0, row=4, columnspan=2)
        
        self.btn_delete.grid(column=0, row=5)
        self.btn_update.grid(column=1, row=5)
        
    def add_book(self):
        title = self.txt_title.get()
        author = self.txt_author.get()
        year = self.txt_year.get()
        if title and author and year:
            self.db.add_book(title, author, year)
            self.refresh_list()
            self.clear_fields()
        else:
            msg.showerror("Input error", "All fields are required.")
    
    def clear_fields(self):
        self.txt_title.delete(0, ttk.END)
        self.txt_author.delete(0, ttk.END)
        self.txt_year.delete(0, ttk.END)
    
    def update_book(self):
        selected_item = self.tv.selection()
        if selected_item:
            book_id = self.tv.item(selected_item, "values")[0]
            if book_id:
                try:
                    book_id = int(book_id)
                    title = self.txt_title.get()
                    author = self.txt_author.get()
                    year = self.txt_year.get()
                    if title and author and year:
                        self.db.update_book(book_id, title, author, year)
                        self.refresh_list()
                        self.clear_fields()
                    else:
                        msg.showerror("Input error", "All fields are required.")
                except ValueError:
                    msg.showerror("Error", "Invalid book ID.")
    
    def delete_book(self):
        selected_item = self.tv.selection()
        if selected_item:
            book_id = self.tv.item(selected_item, "values")[0]
            if book_id:
                try:
                    book_id = int(book_id)
                    self.db.delete_book(book_id)
                    self.refresh_list()
                except ValueError:
                    msg.showerror("Error", "Invalid book ID.")
    
    def refresh_list(self):
        for item in self.tv.get_children():
            self.tv.delete(item)
            
        for row in self.db.get_items():
            self.tv.insert(parent="", index="end", iid=row[0], values=(row[0], row[1], row[2], row[3]))
        
    def export_data(self):
        if not self.tv.get_children():
            msg.showerror("Error", "No data available")
            return
        
        wb = workbook.Workbook()
        ws = wb.active
        
        ws.cell(row=1, column=1, value="Title")
        ws.cell(row=1, column=2, value="Author")
        ws.cell(row=1, column=3, value="Year")
        
        row_count = 2
        for book_id in self.tv.get_children():
            title, author, year = self.tv.item(book_id, "values")[1:]
            ws.cell(row=row_count, column=1, value=title)
            ws.cell(row=row_count, column=2, value=author)
            ws.cell(row=row_count, column=3, value=year)
            row_count += 1
        
        file_path = "exported_books.xlsx"
        wb.save(file_path)
        msg.showinfo("Export Successful", f"Data exported successfully to {file_path}")
        
app = BookManagement()