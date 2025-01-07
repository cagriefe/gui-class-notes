import tkinter as tk
from tkinter import ttk, filedialog as fd, messagebox as msg
from openpyxl import load_workbook, Workbook
import sqlite3
import matplotlib.pyplot as plt
import pandas as pd

class FinanceDatabase:
    def __init__(self, db_name="finance_tracker.db"):
        self.db_name = db_name
        self.create_table()

    def create_table(self):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS finance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                category TEXT NOT NULL,
                amount REAL NOT NULL,
                type TEXT NOT NULL
            );
        """)
        conn.commit()
        conn.close()

    def add_record(self, date, category, amount, type):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("INSERT INTO finance (date, category, amount, type) VALUES (?, ?, ?, ?)", (date, category, amount, type))
        conn.commit()
        conn.close()

    def get_records(self):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("SELECT * FROM finance")
        records = cur.fetchall()
        conn.close()
        return records

    def delete_record(self, record_id):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("DELETE FROM finance WHERE id = ?", (record_id,))
        conn.commit()
        conn.close()

    def update_record(self, record_id, date, category, amount, type):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("UPDATE finance SET date = ?, category = ?, amount = ?, type = ? WHERE id = ?", (date, category, amount, type, record_id))
        conn.commit()
        conn.close()

class FinanceTrackerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Personal Finance Tracker")
        self.geometry("800x400")
        self.db = FinanceDatabase()
        self.create_widgets()
        self.create_layout()
        self.bind("<Control-i>", lambda e: self.import_data())
        self.bind("<Control-e>", lambda e: self.export_data())
        self.bind("<Control-b>", lambda e: self.show_bar_chart())
        self.mainloop()

    def create_widgets(self):
        self.lbl_date = ttk.Label(self, text="Date (YYYY-MM-DD):")
        self.lbl_category = ttk.Label(self, text="Category:")
        self.lbl_amount = ttk.Label(self, text="Amount:")
        self.lbl_type = ttk.Label(self, text="Type:")

        self.txt_date = ttk.Entry(self)
        self.txt_category = ttk.Entry(self)
        self.txt_amount = ttk.Entry(self)
        self.cmb_type = ttk.Combobox(self, values=["Income", "Expense"], state="readonly")

        self.btn_add = ttk.Button(self, text="Add Record", command=self.add_record)
        self.btn_delete = ttk.Button(self, text="Delete Record", command=self.delete_record)

        self.tv = ttk.Treeview(self, columns=("Date", "Category", "Amount", "Type"), show="headings")
        self.tv.heading("Date", text="Date")
        self.tv.heading("Category", text="Category")
        self.tv.heading("Amount", text="Amount")
        self.tv.heading("Type", text="Type")

        self.tv.bind("<<TreeviewSelect>>", self.on_record_select)

    def create_layout(self):
        self.lbl_date.grid(row=0, column=0, padx=10, pady=5)
        self.txt_date.grid(row=0, column=1, padx=10, pady=5)
        self.lbl_category.grid(row=1, column=0, padx=10, pady=5)
        self.txt_category.grid(row=1, column=1, padx=10, pady=5)
        self.lbl_amount.grid(row=2, column=0, padx=10, pady=5)
        self.txt_amount.grid(row=2, column=1, padx=10, pady=5)
        self.lbl_type.grid(row=3, column=0, padx=10, pady=5)
        self.cmb_type.grid(row=3, column=1, padx=10, pady=5)
        self.btn_add.grid(row=4, column=0, columnspan=2, pady=10)
        self.btn_delete.grid(row=5, column=0, columnspan=2, pady=10)
        self.tv.grid(row=6, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

    def add_record(self):
        date = self.txt_date.get()
        category = self.txt_category.get()
        amount = self.txt_amount.get()
        type = self.cmb_type.get()
        if date and category and amount and type:
            try:
                amount = float(amount)
                self.db.add_record(date, category, amount, type)
                self.refresh_records()
            except ValueError:
                msg.showerror("Invalid Input", "Amount must be a number.")
        else:
            msg.showerror("Invalid Input", "All fields are required.")

    def delete_record(self):
        selected_item = self.tv.selection()
        if selected_item:
            record_id = self.tv.item(selected_item)["values"][0]
            self.db.delete_record(record_id)
            self.refresh_records()

    def on_record_select(self, event):
        selected_item = self.tv.selection()
        if selected_item:
            record = self.tv.item(selected_item)["values"]
            self.txt_date.delete(0, tk.END)
            self.txt_date.insert(0, record[1])
            self.txt_category.delete(0, tk.END)
            self.txt_category.insert(0, record[2])
            self.txt_amount.delete(0, tk.END)
            self.txt_amount.insert(0, record[3])
            self.cmb_type.set(record[4])

    def refresh_records(self):
        for item in self.tv.get_children():
            self.tv.delete(item)
        for record in self.db.get_records():
            self.tv.insert("", tk.END, values=record)

    def import_data(self):
        file_path = fd.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = load_workbook(file_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                self.db.add_record(*row)
            self.refresh_records()

    def export_data(self):
        file_path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = Workbook()
            ws = wb.active
            ws.append(["Date", "Category", "Amount", "Type"])
            for record in self.db.get_records():
                ws.append(record[1:])
            wb.save(file_path)

    def show_bar_chart(self):
        records = self.db.get_records()
        df = pd.DataFrame(records, columns=["ID", "Date", "Category", "Amount", "Type"])
        df["Date"] = pd.to_datetime(df["Date"])
        df["Month"] = df["Date"].dt.to_period("M")
        income = df[df["Type"] == "Income"].groupby("Month")["Amount"].sum()
        expense = df[df["Type"] == "Expense"].groupby("Month")["Amount"].sum()
        plt.bar(income.index.astype(str), income, label="Income")
        plt.bar(expense.index.astype(str), expense, label="Expense", bottom=income)
        plt.xlabel("Month")
        plt.ylabel("Amount")
        plt.title("Monthly Income and Expenses")
        plt.legend()
        plt.show()

if __name__ == "__main__":
    app = FinanceTrackerApp()