# Final Exam Question: Personal Finance Tracker Application
# Scenario:
# You are tasked with creating a Personal Finance Tracker application that allows users to manage their income and expenses. The application should provide functionalities to add, view, update, and delete financial records. Additionally, the application should support importing records from an Excel file and exporting the current records to an Excel file. The application should also display a bar chart showing the total income and expenses for each month.

# Requirements:
# Database Setup:

# Create a SQLite database named finance_tracker.db.
# Create a table named finance with the following columns:
# id (integer, primary key, autoincrement)
# date (text, not null)
# category (text, not null)
# amount (real, not null)
# type (text, not null, values: 'Income' or 'Expense')
# GUI Application:

# Use tkinter and ttk for the GUI components.
# The main window should have the following components:
# An entry widget to input the date.
# An entry widget to input the category.
# An entry widget to input the amount.
# A combobox to select the type (Income or Expense).
# A button to add the record to the database.
# A Treeview to display the list of records with columns for Date, Category, Amount, and Type.
# A button to delete the selected record.
# Menu options to import records from an Excel file and export records to an Excel file.
# A keyboard shortcut to display a bar chart of total income and expenses for each month.
# Functionality:

# Add Record: Add a new financial record to the database.
# View Records: Display all records in the Treeview.
# Update Record: Allow updating the selected record.
# Delete Record: Delete the selected record from the database.
# Import Records: Import records from an Excel file. The file will have columns: Date, Category, Amount, and Type.
# Export Records: Export the current records to an Excel file with columns for Date, Category, Amount, and Type.
# Bar Chart: Display a bar chart showing the total income and expenses for each month.
# Error Handling:

# Ensure that all input fields are filled before adding a record to the database.
# Handle any database connection errors gracefully.
# Validate the format of the imported Excel file.
# Expected Outputs:
# A functional Personal Finance Tracker application with the described features.
# An Excel file named exported_finance.xlsx containing the current records when the export functionality is used.
# A bar chart displaying the total income and expenses for each month.

import tkinter as tk
import ttkbootstrap as ttk
from tkinter import ttk, filedialog as fd, messagebox as msg
from openpyxl import load_workbook, Workbook
import sqlite3
import matplotlib.pyplot as plt
import pandas as pd

class FinanceDatabase:
    def __init__(self, db_name="finance_tracker.db"):
        self.db_name = db_name

    def create_table(self):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("""
                    CREATE TABLE IF NOT EXISTS finance (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        date TEXT NOT NULL,
                        category TEXT NOT NULL,
                        amount REAL NOT NULL,
                        type TEXT NOT NULL
                    )
                    """)
        conn.commit()
        conn.close()

    def add_record(self, date, category, amount, type):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("INSERT INTO finance (date, category, amount, type) VALUES (?, ?, ?, ?)",
                    (date, category, amount, type))
        conn.commit()
        conn.close()

    def get_records(self):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("SELECT * FROM finance")
        records = cur.fetchall()
        conn.close()
        return records

    def delete_record(self, record_id):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("DELETE FROM finance WHERE id = ?", (record_id,))
        conn.commit()
        conn.close()

    def update_record(self, record_id, date, category, amount, type):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("UPDATE finance SET date = ?, category = ?, amount = ?, type = ? WHERE id = ?",
                    (date, category, amount, type, record_id))
        conn.commit()
        conn.close()

class FinanceTrackerApp(ttk.Window):
    def __init__(self):
        super().__init__()
        self.geometry("800x400")
        self.title("Final")
        self.db = FinanceDatabase()
        self.create_widgets()
        self.create_layout()
        

    def create_widgets(self):
        self.date_lbl = ttk.Label(self, text="Date:")
        self.date_entry = ttk.Entry(self)
        
        self.category_lbl = ttk.Label(self, text="Category:")
        self.category_entry = ttk.Entry(self)
        
        self.amount_lbl = ttk.Label(self, text="Amount:")
        self.amount_entry = ttk.Entry(self)
        
        self.type_lbl = ttk.Label(self, text="Type:")
        self.type_combo = ttk.Combobox(self, values=["Income", "Expense"])
        
        self.btn_save = ttk.Button(self, text="Save", width=50, command=self.add_record)
        self.btn_delete = ttk.Button(self, text="Delete", width=50, command=self.delete_record)
        
        self.tv = ttk.Treeview(self, height=10, show="headings")
        self.tv["columns"] = ("date", "category", "amount", "type")
        self.tv["selectmode"] = "browse"
        self.tv_scroll = ttk.Scrollbar(self, orient="vertical", command=self.tv.yview)
        self.tv.configure(yscrollcommand=self.tv_scroll.set)
        
        
        

    def create_layout(self):
        self.columnconfigure(index=0, weight=1, uniform="eq")
        self.columnconfigure(index=1, weight=1, uniform="eq")
        
        self.rowconfigure(index=0, weight=1, uniform="eq")
        self.rowconfigure(index=1, weight=1, uniform="eq")
        self.rowconfigure(index=2, weight=1, uniform="eq")
        self.rowconfigure(index=3, weight=1, uniform="eq")
        self.rowconfigure(index=4, weight=2, uniform="eq")
        self.rowconfigure(index=5, weight=2, uniform="eq")
        
        
        self.date_lbl.grid(column=0, row=0)
        self.date_entry.grid(column=1, row=0)

        self.category_lbl.grid(column=0, row=1)
        self.category_entry.grid(column=1, row=1)
        
        self.amount_lbl.grid(column=0, row=2)
        self.amount_entry.grid(column=1, row=2)
        
        self.type_lbl.grid(column=0, row=3)
        self.type_combo.grid(column=1, row=3)
        
        self.btn_save.grid(column=0, row=4)
        
        self.btn_delete.grid(column=0, row=5)
        
    def add_record(self):
        date = self.date_entry.get()
        category = self.category_entry.get()
        amount = self.amount_entry.get()
        if date and category and amount:
            self.db.add_record(date=date, category=category, amount=amount)
            self.refresh_records()
        

    def delete_record(self):
        selected_item = self.tv.selection()
        if selected_item:
            item_id = int(selected_item[0])
            self.db.delete_record(item_id)
            self.refresh_records()

    def on_record_select(self, event):
        selected_item = self.tv.selection()
        if selected_item:
            item_id = int(selected_item[0])
            return item_id

    def refresh_records(self):
        for item in self.tv.get_children():
            self.tv.delete(item)
            
        for row in self.db.get_records():
            self.tv.insert(parent="", index="end", iid=row[0], values=(row[1], row[2], row[3]))

    def import_data(self):
        file_filter = (["Excel Files", "*.xslsx"],
                       ["All files", "*.*"])
        
        source_file = fd.askopenfile(title="Choose File", filetypes=file_filter)
        
        if source_file is not None:
            wb = load_workbook(source_file.name)
            ws1 = wb.active
            skip_first = True
            
            for row in ws1.iter_rows(values_only=True):
                if skip_first:
                    skip_first = False
                    continue
                
                item = row[0], row[1], row[2], row[3]
                
                if item:
                    self.db.add_record(item)
                    
            self.refresh_records()

    def export_data(self):
        if not self.tv.get_children():
            msg.showerror(title="Error", message="No data to export")
            return
        
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        headers = ["Date", "Category", "Amount", "Type"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Add data
        for row_num, item in enumerate(self.tv.get_children(), 2):
            row_data = self.tv.item(item, "values")
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Save file
        file_path = fd.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path:
            wb.save(file_path)
            msg.showinfo(title="Success", message="Data exported successfully")

    def show_bar_chart(self):
        items = self.db.get_records()
        

# Database setup and application launch
finance_db = FinanceDatabase()
finance_db.create_table()
app = FinanceTrackerApp()