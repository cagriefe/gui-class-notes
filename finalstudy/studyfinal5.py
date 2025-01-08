import ttkbootstrap as ttk  # Importing ttkbootstrap for themed tkinter widgets
from tkinter import filedialog as fd  # Importing filedialog for file operations
from ttkbootstrap.dialogs import Messagebox as msg  # Importing Messagebox for displaying messages
from openpyxl import load_workbook  # Importing load_workbook for reading Excel files
from openpyxl import Workbook  # Importing Workbook for creating Excel files
import sqlite3
import matplotlib.pyplot as plt

class ToDoDatabase:

    def __init__(self, db_name="todo_list.db"):
        self.db_name = db_name

    def create_table(self):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("""
            create table if not exists todolist (
                id integer primary key autoincrement,
                item text not null,
                status text default 'Pending'
            );
        """)
        conn.commit()
        conn.close()

    def add_item(self, item):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("insert into todolist (item) values (?)", (item,))
        conn.commit()
        conn.close()

    def get_items(self):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("select id, item, status from todolist")
        items = cur.fetchall()
        conn.close()

        return items

    def delete_item(self, item_id):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("delete from todolist where id = ?", (item_id,))
        conn.commit()
        conn.close()

    def update_status(self, item_id, status):
        conn = sqlite3.connect(self.db_name)
        cur = conn.cursor()
        cur.execute("update todolist set status = ? where id = ?", 
                    (status, item_id))
        conn.commit()
        conn.close()


class ToDoList(ttk.Window):  # Defining a class that inherits from ttk.Window
    def __init__(self):  # Constructor method
        super().__init__()
        self.geometry("600x720")
        self.title("To-do list app")
        self.db = ToDoDatabase()
        self.status_var = ttk.IntVar()
        self.create_widgets()
        self.create_layout()
        self.mainloop()
        
    def create_widgets(self):  # Method to create widgets
        self.frm_add = ttk.LabelFrame(self, text="Add todo item")
        self.entry_item = ttk.Entry(self.frm_add)
        self.btn_add = ttk.Button(self.frm_add, text="Add", command=self.add_item)
        
        self.frm_list = ttk.LabelFrame(self, text="Todo list")
        self.tv = ttk.Treeview(self.frm_list, columns=("Item","Status"), show="headings", selectmode="browse")
        self.tv.heading("Item", text="Item")
        self.tv.heading("Status", text="Status")
        
        self.frm_action = ttk.LabelFrame(self, text="Actions")
        self.btn_delete = ttk.Button(self.frm_action, text="Delete", command=self.delete_item)
        self.cbtn_status = ttk.Checkbutton(self.frm_action, text="Mark as Completed", variable=self.status_var, 
                                           command=self.toggle_status, bootstyle="round-toggle")
        
        self.bind("<<TreeviewSelect>>", self.on_item_select)
        self.entry_item.bind("<Return>", lambda e: self.add_item())
        self.bind("<Control-i>", lambda e : self.import_data())
        self.bind("<Control-e>", lambda e : self.export_data())
        self.bind("<Control-Shift-C>", lambda e : self.show_pie_chart())
        
        self.refresh_list()
        
    def create_layout(self):  # Method to create layout
        self.frm_add.pack(fill="x", padx=15, pady=(15,0))
        self.frm_list.pack(fill="both", expand=True, padx=15, pady=(15,0))
        self.frm_action.pack(fill="x", padx=15, pady=15)
        
        self.entry_item.pack(side="left", padx=(15,0), pady=15, fill="x", expand=True)
        self.btn_add.pack(side="right", padx=15, pady=15)
        
        self.tv.pack(fill="both", expand=True, padx=15, pady=15)
        
        self.btn_delete.pack(side="left", padx=(15,0), pady=15)
        self.cbtn_status.pack(side="right", padx=(15,0), pady=15)
        
    def enable_actions(self):  # Method to enable action buttons
        self.btn_delete.configure(state="normal")
        self.cbtn_status.configure(state="normal")
        
    def disable_actions(self):  # Method to disable action buttons
        self.btn_delete.configure(state="disabled")
        self.cbtn_status.configure(state="disabled")
        
    def on_item_select(self, event):  # Method to handle item selection
        selected_item = self.tv.selection()
        if selected_item():
            item_id = int(selected_item[0])
            item_status = self.tv.item(item_id)["values"][1]
            if item_status == "Done":
                self.status_var.set(1)
            else:
                self.status_var.set(0)

    def refresh_list(self):  # Method to refresh the list
        for item in self.tv.get_children():  # Clearing the Treeview content
            self.tv.delete(item)  # Deleting each item

        for row in self.db.get_items():  # Fetching updated data and populating the Treeview
            self.tv.insert(parent="", index="end", iid=row[0], values=(row[1], row[2]))  
                                                                    # Inserting each row
        tv_items = self.tv.get_children()  # Getting the Treeview items
        if tv_items:  # If there are items in the Treeview
            self.enable_actions()  # Enabling the action buttons
            self.tv.selection_set(tv_items[-1])  # Selecting the last item
        else:  # If there are no items in the Treeview
            self.disable_actions()  # Disabling the action buttons
            
    def add_item(self):  # Method to add an item
        item = self.entry_item.get().strip()
        if item:
            self.db.add_item(item)
            self.refresh_list()
            self.entry_item.delete(0, "end")
            self.entry_item.focus_set()

    def delete_item(self):  # Method to delete an item
        selected_item = self.tv.selection()
        if selected_item:
            item_id = int(selected_item[0])
            self.db.delete_item(item_id)
            self.refresh_list()
            
    def toggle_status(self):  # Method to toggle the status of an item
        selected_item = self.tv.selection()
        if selected_item:
            item_id = int(selected_item[0])
            if self.status_var.get() == 1:
                new_status = "Done"
            else:
                new_status = "Pending"
            self.db.update_status(item_id, new_status)
            self.refresh_list()
            self.tv.selection_set(selected_item)
            
    def import_data(self):  # Method to import data
        file_filter = (("Excel Files", "*.xlsx"), ("All files","*.*"))
        source_file = fd.askopenfile(title="Choose file", filetypes=file_filter)
        
        if source_file is not None:
            wb = load_workbook(source_file.name)
            ws1 = wb.active
            skip_first = True
            
            for row in ws1.iter_rows(values_only=True):
                if skip_first:
                    skip_first = False
                    continue
                item, status = row[0], row[1]
                if item:
                    self.db.add_item(item)
                    
                    if status == "Done":
                        self.db.update_status(self.db.get_items()[-1][0],"Done")
            
            self.refresh_list()
        

    def export_data(self):  # Method to export data
        if not self.tv.get_children():  # If there are no items in the Treeview
            msg.show_error(title="Export Data", message="No data available")  # Showing an error message
            return

        wb = Workbook()  # Creating a new workbook
        ws = wb.active  # Getting the active worksheet

        ws.cell(row=1, column=1, value="Item")  # Adding column header for Item
        ws.cell(row=1, column=2, value="Status")  # Adding column header for Status

        row_count = 2  # Starting row count for data
        for item_id in self.tv.get_children():  # Reading data from Treeview
            ws.cell(row=row_count, column=1,
                    value=self.tv.item(item_id)["values"][0])  # Adding item to the worksheet
            ws.cell(row=row_count, column=2, 
                    value=self.tv.item(item_id)["values"][1])  # Adding status to the worksheet
            row_count += 1  # Incrementing row count

        total_items_row = row_count + 1  # Row for total items
        completed_items_row = row_count + 2  # Row for completed items
        not_completed_items_row = row_count + 3  # Row for not completed items

        ws.cell(row=total_items_row, column=1, value="Total items:")  # Adding label for total items
        ws.cell(row=total_items_row, column=2, 
                value=f"=COUNTA(A2:A{row_count - 1})")  # Adding formula for total items

        ws.cell(row=completed_items_row, column=1,
                value="Completed items:")  # Adding label for completed items
        ws.cell(row=completed_items_row, column=2, 
                value=f"=COUNTIF(B2:B{row_count - 1}, \"Done\")")  # Adding formula for completed items

        ws.cell(row=not_completed_items_row, column=1, 
                value="Not completed items:")  # Adding label for not completed items
        ws.cell(row=not_completed_items_row, column=2, 
                value=f"=COUNTIF(B2:B{row_count - 1}, \"Pending\")")  # Adding formula for not completed items

        wb.save("exported_list.xlsx")  # Saving the workbook
        msg.show_info(title="Export Data",
                      message="exported and saved as 'exported_list.xlsx'.")  # Showing an info message
    
        
    def show_pie_chart(self):
        items = self.db.get_items()
        if len(items) == 0:
            return
        
        completed = sum(1 for item in items if item[2]=="Done")
        not_completed = sum(1 for item in items if item[2]=="Pending")
        
        labels = ["Completed", "Not Completed"]
        
        sizes = [completed, not_completed]
        
        colors = ["lightblue", "red"]
        
        explode = [0.1,0]
        
        plt.pie(sizes, explode=explode, labels=labels, colors=colors, autopct="%.1f%%", shadow=True, startangle=90)
        plt.title("Todo list Status")
        plt.show()
        
        
gradebook_db = ToDoDatabase()
gradebook_db.create_table()

app = ToDoList()  # Creating an instance of ToDoList

